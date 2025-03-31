import asyncio
import json
import logging
from aiogram import Bot, Dispatcher, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, InputMediaPhoto, InputMediaDocument
from aiogram.filters import Command
from aiogram.enums import ChatType
from config import Config, load_config
import aiohttp
from io import BytesIO
from openpyxl import load_workbook

# ÐÐ°ÑÑ‚Ñ€Ð¾Ð¹ÐºÐ° Ð»Ð¾Ð³Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð¸Ñ Ð´Ð»Ñ Ð²Ñ‹Ð²Ð¾Ð´Ð° Ð¾Ñ‚Ð»Ð°Ð´Ð¾Ñ‡Ð½Ð¾Ð¹ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ð¸
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Ð—Ð°Ð³Ñ€ÑƒÐ·ÐºÐ° ÐºÐ¾Ð½Ñ„Ð¸Ð³ÑƒÑ€Ð°Ñ†Ð¸Ð¸ Ð±Ð¾Ñ‚Ð° Ð¸Ð· Ñ„Ð°Ð¹Ð»Ð° config.py
config: Config = load_config()
BOT_TOKEN: str = config.tg_bot.token

# Ð˜Ð½Ð¸Ñ†Ð¸Ð°Ð»Ð¸Ð·Ð°Ñ†Ð¸Ñ Ð±Ð¾Ñ‚Ð° Ð¸ Ð´Ð¸ÑÐ¿ÐµÑ‚Ñ‡ÐµÑ€Ð°
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# Ð“Ð»Ð¾Ð±Ð°Ð»ÑŒÐ½Ñ‹Ðµ Ð¿ÐµÑ€ÐµÐ¼ÐµÐ½Ð½Ñ‹Ðµ
media_groups = {}
faq = {}
YANDEX_DISK_TOKEN = "y0__xD-s5TpBxijkzYgie2UyhKmZIBRVpLHIieiT1CMAYGOMXpgHQ"  # ÐŸÑ€Ð¾Ð²ÐµÑ€ÑŒÑ‚Ðµ Ð¿Ð¾Ð»Ð½Ñ‹Ð¹ Ñ‚Ð¾ÐºÐµÐ½
FILE_PATH = "faq.xlsx"
MANAGER_CHAT_ID = -4634857148

# Ð¤ÑƒÐ½ÐºÑ†Ð¸Ñ Ð·Ð°Ð³Ñ€ÑƒÐ·ÐºÐ¸ Ñ‚ÐµÐºÑƒÑ‰ÐµÐ³Ð¾ ID Ð°Ð½ÐºÐµÑ‚Ñ‹ Ð¸Ð· Ñ„Ð°Ð¹Ð»Ð°
def load_survey_id():
    """Ð§Ð¸Ñ‚Ð°ÐµÑ‚ Ð¿Ð¾ÑÐ»ÐµÐ´Ð½Ð¸Ð¹ ID Ð°Ð½ÐºÐµÑ‚Ñ‹ Ð¸Ð· Ñ„Ð°Ð¹Ð»Ð° survey_id.txt. Ð•ÑÐ»Ð¸ Ñ„Ð°Ð¹Ð»Ð° Ð½ÐµÑ‚, Ð²Ð¾Ð·Ð²Ñ€Ð°Ñ‰Ð°ÐµÑ‚ 1."""
    try:
        with open("survey_id.txt", "r") as f:
            return int(f.read())
    except:
        return 1

# Ð¤ÑƒÐ½ÐºÑ†Ð¸Ñ ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½Ð¸Ñ Ñ‚ÐµÐºÑƒÑ‰ÐµÐ³Ð¾ ID Ð°Ð½ÐºÐµÑ‚Ñ‹ Ð² Ñ„Ð°Ð¹Ð»
def save_survey_id(counter):
    """Ð—Ð°Ð¿Ð¸ÑÑ‹Ð²Ð°ÐµÑ‚ Ð½Ð¾Ð²Ñ‹Ð¹ ID Ð°Ð½ÐºÐµÑ‚Ñ‹ Ð² Ñ„Ð°Ð¹Ð» survey_id.txt."""
    with open("survey_id.txt", "w") as f:
        f.write(str(counter))

survey_id_counter = load_survey_id()

# Ð¤ÑƒÐ½ÐºÑ†Ð¸Ñ Ð·Ð°Ð³Ñ€ÑƒÐ·ÐºÐ¸ Ð´Ð°Ð½Ð½Ñ‹Ñ… Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»ÐµÐ¹ Ð¸Ð· Ñ„Ð°Ð¹Ð»Ð°
def load_users():
    """Ð§Ð¸Ñ‚Ð°ÐµÑ‚ Ð´Ð°Ð½Ð½Ñ‹Ðµ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»ÐµÐ¹ Ð¸Ð· users.json. Ð•ÑÐ»Ð¸ Ñ„Ð°Ð¹Ð»Ð° Ð½ÐµÑ‚, Ð²Ð¾Ð·Ð²Ñ€Ð°Ñ‰Ð°ÐµÑ‚ Ð¿ÑƒÑÑ‚Ð¾Ð¹ ÑÐ»Ð¾Ð²Ð°Ñ€ÑŒ."""
    try:
        with open("users.json", "r") as f:
            return json.load(f)
    except:
        return {}

# Ð¤ÑƒÐ½ÐºÑ†Ð¸Ñ ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½Ð¸Ñ Ð´Ð°Ð½Ð½Ñ‹Ñ… Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»ÐµÐ¹ Ð² Ñ„Ð°Ð¹Ð»
def save_users(users):
    """Ð—Ð°Ð¿Ð¸ÑÑ‹Ð²Ð°ÐµÑ‚ Ð´Ð°Ð½Ð½Ñ‹Ðµ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»ÐµÐ¹ Ð² users.json Ñ Ð¿Ð¾Ð´Ð´ÐµÑ€Ð¶ÐºÐ¾Ð¹ ÐºÐ¸Ñ€Ð¸Ð»Ð»Ð¸Ñ†Ñ‹."""
    with open("users.json", "w") as f:
        json.dump(users, f, ensure_ascii=False, indent=4)

users = load_users()

# Ð¡Ð¿Ð¸ÑÐ¾Ðº Ð²Ð¾Ð¿Ñ€Ð¾ÑÐ¾Ð² Ð°Ð½ÐºÐµÑ‚Ñ‹
questions = [
    "Ð’Ð°ÑˆÐµ Ð¸Ð¼Ñ Ð¸ Ñ„Ð°Ð¼Ð¸Ð»Ð¸Ñ",
    "Ð’Ð°Ñˆ Ð½Ð¸Ðº Ð² Telegram (Ñ‡ÐµÑ€ÐµÐ· @)",
    "ÐÐ°Ð¿Ð¸ÑˆÐ¸Ñ‚Ðµ ÐºÐ¾Ð½Ñ‚Ð°ÐºÑ‚Ð½Ñ‹Ð¹ Ñ‚ÐµÐ»ÐµÑ„Ð¾Ð½ Ð´Ð»Ñ ÑÐ²ÑÐ·Ð¸",
    "ÐÐ°Ð¿Ð¸ÑˆÐ¸Ñ‚Ðµ Ð½Ð°Ð¸Ð¼ÐµÐ½Ð¾Ð²Ð°Ð½Ð¸Ðµ Ñ‚Ð¾Ð²Ð°Ñ€Ð°",
    "Ð’ÑÑ‚Ð°Ð²ÑŒÑ‚Ðµ ÑÑÑ‹Ð»ÐºÑƒ Ð½Ð° Ñ‚Ð¾Ð²Ð°Ñ€ Ð½Ð° Ð¼Ð°Ñ€ÐºÐµÑ‚Ð¿Ð»ÐµÐ¹ÑÐ°Ñ… (ÐµÑÐ»Ð¸ ÐµÑÑ‚ÑŒ)",
    "ÐšÐ°ÐºÐ¾Ðµ ÐºÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾ Ð²Ð°Ð¼ Ð½ÑƒÐ¶Ð½Ð¾?",
    "ÐŸÑ€Ð¸ÐºÑ€ÐµÐ¿Ð¸Ñ‚Ðµ Ñ„Ð¾Ñ‚Ð¾ Ñ‚Ð¾Ð²Ð°Ñ€Ð° (Ð¸Ð»Ð¸ PDF, Excel-Ñ„Ð°Ð¹Ð»). Ð•ÑÐ»Ð¸ Ñ„Ð°Ð¹Ð»Ð¾Ð² Ð½ÐµÑÐºÐ¾Ð»ÑŒÐºÐ¾, Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÑŒÑ‚Ðµ Ð¸Ñ… Ð¾Ð´Ð½Ð¸Ð¼ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸ÐµÐ¼ Ð¸Ð»Ð¸ Ð¿Ð¾ Ð¾Ð´Ð½Ð¾Ð¼Ñƒ, Ð·Ð°Ñ‚ÐµÐ¼ Ð½Ð°Ð¿Ð¸ÑˆÐ¸Ñ‚Ðµ 'Ð“Ð¾Ñ‚Ð¾Ð²Ð¾'",
    "ÐÐ°Ð¿Ð¸ÑˆÐ¸Ñ‚Ðµ Ñ€Ð°Ð·Ð¼ÐµÑ€Ñ‹ Ñ‚Ð¾Ð²Ð°Ñ€Ð° Ð¸ ÐºÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾ ÐºÐ°Ð¶Ð´Ð¾Ð³Ð¾ Ñ€Ð°Ð·Ð¼ÐµÑ€Ð°",
    "Ð£ÐºÐ°Ð¶Ð¸Ñ‚Ðµ Ñ†Ð²ÐµÑ‚ Ñ‚Ð¾Ð²Ð°Ñ€Ð°",
    "Ð£ÐºÐ°Ð¶Ð¸Ñ‚Ðµ, Ð½ÑƒÐ¶Ð½Ñ‹ Ð»Ð¸ Ð´Ð¾Ð¿Ð¾Ð»Ð½Ð¸Ñ‚ÐµÐ»ÑŒÐ½Ñ‹Ðµ ÑÐ»ÐµÐ¼ÐµÐ½Ñ‚Ñ‹ (Ð½Ð°Ð¿Ñ€Ð¸Ð¼ÐµÑ€, Ð°ÐºÑÐµÑÑÑƒÐ°Ñ€Ñ‹, Ð¸Ð½ÑÑ‚Ñ€ÑƒÐ¼ÐµÐ½Ñ‚Ñ‹, Ð²Ð°Ð¶Ð½Ñ‹Ðµ Ð´ÐµÑ‚Ð°Ð»Ð¸).",
    "ÐšÐ°ÐºÐ°Ñ ÑƒÐ¿Ð°ÐºÐ¾Ð²ÐºÐ° Ð½ÑƒÐ¶Ð½Ð°? (ÐµÑÑ‚ÑŒ Ð»Ð¸ Ð¾ÑÐ¾Ð±ÐµÐ½Ð½Ð¾ÑÑ‚Ð¸ ÑƒÐ¿Ð°ÐºÐ¾Ð²ÐºÐ¸ Ð¸ Ð´Ð¾Ð¿Ð¾Ð»Ð½Ð¸Ñ‚ÐµÐ»ÑŒÐ½Ð¾Ð¹ Ð·Ð°Ñ‰Ð¸Ñ‚Ñ‹?)",
    "Ð‘Ñ€ÐµÐ½Ð´Ð¸Ð½Ð³ (Ð½ÑƒÐ¶Ð½Ð¾ Ð»Ð¸ Ð±Ñ€ÐµÐ½Ð´Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð¸Ðµ Ñ‚Ð¾Ð²Ð°Ñ€Ð°, ÐµÑÐ»Ð¸ Ð´Ð°, ÑƒÐºÐ°Ð¶Ð¸Ñ‚Ðµ Ð¼ÐµÑÑ‚Ð¾ Ð¸ Ñ€Ð°Ð·Ð¼ÐµÑ€)",
    "Ð’Ñ‹Ð±ÐµÑ€Ð¸Ñ‚Ðµ ÑÑ€Ð¾Ðº Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸ (10-13, 15-18, 25-30 Ð´Ð½ÐµÐ¹, ÐÐ²Ð¸Ð°)",
    "Ð•ÑÑ‚ÑŒ Ð»Ð¸ Ð´Ð¾Ð¿Ð¾Ð»Ð½Ð¸Ñ‚ÐµÐ»ÑŒÐ½Ñ‹Ðµ ÑƒÑ‚Ð¾Ñ‡Ð½ÐµÐ½Ð¸Ñ?",
    "ÐŸÐµÑ€ÐµÑ‡Ð¸ÑÐ»Ð¸Ñ‚Ðµ Ð²Ð¾Ð¿Ñ€Ð¾ÑÑ‹ Ð´Ð»Ñ Ð¿Ð¾ÑÑ‚Ð°Ð²Ñ‰Ð¸ÐºÐ° (ÐµÑÐ»Ð¸ ÐµÑÑ‚ÑŒ, ÑƒÐºÐ°Ð¶Ð¸Ñ‚Ðµ Ð·Ð´ÐµÑÑŒ)"
]

user_answers = {}

# Ð¤ÑƒÐ½ÐºÑ†Ð¸Ñ Ð·Ð°Ð³Ñ€ÑƒÐ·ÐºÐ¸ FAQ Ð¸Ð· Ñ„Ð°Ð¹Ð»Ð° Ð½Ð° Ð¯Ð½Ð´ÐµÐºÑ.Ð”Ð¸ÑÐºÐµ
async def load_faq_from_yandex_disk():
    """Ð—Ð°Ð³Ñ€ÑƒÐ¶Ð°ÐµÑ‚ Ñ„Ð°Ð¹Ð» faq.xlsx Ñ Ð¯Ð½Ð´ÐµÐºÑ.Ð”Ð¸ÑÐºÐ° Ð¸ Ð¿Ð°Ñ€ÑÐ¸Ñ‚ ÐµÐ³Ð¾ Ð² ÑÐ»Ð¾Ð²Ð°Ñ€ÑŒ FAQ."""
    url = f"https://cloud-api.yandex.net/v1/disk/resources/download?path=/{FILE_PATH}"
    headers = {"Authorization": f"OAuth {YANDEX_DISK_TOKEN}"}
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers) as response:
                if response.status == 200:
                    download_data = await response.json()
                    download_url = download_data["href"]
                    async with session.get(download_url) as file_response:
                        if file_response.status == 200:
                            xlsx_content = await file_response.read()
                            faq_dict = {}
                            workbook = load_workbook(filename=BytesIO(xlsx_content))
                            sheet = workbook.active
                            for row in sheet.iter_rows(min_row=2, values_only=True):
                                question, answer = row[0], row[1]
                                logger.debug(f"ÐžÐ±Ñ€Ð°Ð±Ð°Ñ‚Ñ‹Ð²Ð°ÐµÐ¼ ÑÑ‚Ñ€Ð¾ÐºÑƒ: {question} | {answer}")
                                if question and answer:
                                    faq_dict[str(question).lower()] = str(answer)
                            logger.debug(f"FAQ Ð·Ð°Ð³Ñ€ÑƒÐ¶ÐµÐ½: {faq_dict}")
                            return faq_dict
                        else:
                            logger.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð·Ð°Ð³Ñ€ÑƒÐ·ÐºÐ¸ Ñ„Ð°Ð¹Ð»Ð°: {file_response.status}")
                            return {}
                else:
                    logger.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ð¸Ñ ÑÑÑ‹Ð»ÐºÐ¸: {response.status} - {await response.text()}")
                    return {}
    except Exception as e:
        logger.error(f"Ð˜ÑÐºÐ»ÑŽÑ‡ÐµÐ½Ð¸Ðµ Ð¿Ñ€Ð¸ Ð·Ð°Ð³Ñ€ÑƒÐ·ÐºÐµ FAQ: {str(e)}")
        return {}

# Ð¤ÑƒÐ½ÐºÑ†Ð¸Ñ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð¸Ñ Ð³Ð»Ð¾Ð±Ð°Ð»ÑŒÐ½Ð¾Ð³Ð¾ ÑÐ»Ð¾Ð²Ð°Ñ€Ñ FAQ
async def update_faq():
    """ÐžÐ±Ð½Ð¾Ð²Ð»ÑÐµÑ‚ Ð³Ð»Ð¾Ð±Ð°Ð»ÑŒÐ½Ñ‹Ð¹ ÑÐ»Ð¾Ð²Ð°Ñ€ÑŒ faq Ð´Ð°Ð½Ð½Ñ‹Ð¼Ð¸ Ð¸Ð· Ð¯Ð½Ð´ÐµÐºÑ.Ð”Ð¸ÑÐºÐ°, Ð´Ð¾Ð±Ð°Ð²Ð»ÑÐµÑ‚ Ð·Ð°Ð³Ð»ÑƒÑˆÐºÑƒ Ð¿Ñ€Ð¸ Ð¾ÑˆÐ¸Ð±ÐºÐµ."""
    global faq
    faq = await load_faq_from_yandex_disk()
    if not faq:
        logger.warning("FAQ Ð¿ÑƒÑÑ‚Ð¾Ð¹, Ð¸ÑÐ¿Ð¾Ð»ÑŒÐ·ÑƒÑŽ Ð·Ð°Ð³Ð»ÑƒÑˆÐºÑƒ")
        faq = {"Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ°": "Ð˜Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ñ Ð¾ Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐµ Ð²Ñ€ÐµÐ¼ÐµÐ½Ð½Ð¾ Ð½ÐµÐ´Ð¾ÑÑ‚ÑƒÐ¿Ð½Ð°"}

# Ð¤ÑƒÐ½ÐºÑ†Ð¸Ñ Ð¿ÐµÑ€Ð¸Ð¾Ð´Ð¸Ñ‡ÐµÑÐºÐ¾Ð³Ð¾ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð¸Ñ FAQ
async def periodic_faq_update():
    """Ð—Ð°Ð¿ÑƒÑÐºÐ°ÐµÑ‚ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð¸Ðµ FAQ ÐºÐ°Ð¶Ð´Ñ‹Ðµ 60 Ð¼Ð¸Ð½ÑƒÑ‚ Ð² Ñ„Ð¾Ð½Ð¾Ð²Ð¾Ð¼ Ñ€ÐµÐ¶Ð¸Ð¼Ðµ."""
    while True:
        await update_faq()
        logger.debug("FAQ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½ Ð¸Ð· Ð¯Ð½Ð´ÐµÐºÑ.Ð”Ð¸ÑÐºÐ°")
        await asyncio.sleep(3600)

# ÐšÐ»Ð°Ð²Ð¸Ð°Ñ‚ÑƒÑ€Ð° Ð³Ð»Ð°Ð²Ð½Ð¾Ð³Ð¾ Ð¼ÐµÐ½ÑŽ
start_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Ð—Ð°Ð¿Ð¾Ð»Ð½Ð¸Ñ‚ÑŒ Ð°Ð½ÐºÐµÑ‚Ñƒ")],
        [KeyboardButton(text="Ð§Ð°ÑÑ‚Ñ‹Ðµ Ð²Ð¾Ð¿Ñ€Ð¾ÑÑ‹")],
        [KeyboardButton(text="ÐÐ°Ð·Ð°Ð´")]
    ],
    resize_keyboard=True
)

# Ð˜Ð½Ð»Ð°Ð¹Ð½-ÐºÐ»Ð°Ð²Ð¸Ð°Ñ‚ÑƒÑ€Ð° Ð´Ð»Ñ ÑÐ¾Ð³Ð»Ð°ÑÐ¸Ñ Ð½Ð° Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÑƒ Ð´Ð°Ð½Ð½Ñ‹Ñ…
consent_keyboard = InlineKeyboardMarkup(
    inline_keyboard=[
        [InlineKeyboardButton(text="âœ… Ð”Ð°", callback_data="consent_yes")],
        [InlineKeyboardButton(text="âŒ ÐÐµÑ‚", callback_data="consent_no")],
        [InlineKeyboardButton(text="ðŸ“„ ÐŸÐ¾ÑÐ¼Ð¾Ñ‚Ñ€ÐµÑ‚ÑŒ Ð¾Ñ„ÐµÑ€Ñ‚Ñƒ", callback_data="view_offer")]
    ]
)

# Ð˜Ð½Ð»Ð°Ð¹Ð½-ÐºÐ»Ð°Ð²Ð¸Ð°Ñ‚ÑƒÑ€Ð° Ð´Ð»Ñ Ð²Ñ‚Ð¾Ñ€Ð¾Ð³Ð¾ ÑˆÐ°Ð½ÑÐ° ÑÐ¾Ð³Ð»Ð°ÑÐ¸Ñ
consent_second_keyboard = InlineKeyboardMarkup(
    inline_keyboard=[
        [InlineKeyboardButton(text="âœ… Ð”Ð°", callback_data="consent_yes")],
        [InlineKeyboardButton(text="âŒ ÐÐµÑ‚", callback_data="final_no")]
    ]
)

# Ð˜Ð½Ð»Ð°Ð¹Ð½-ÐºÐ»Ð°Ð²Ð¸Ð°Ñ‚ÑƒÑ€Ð° Ð´Ð»Ñ Ð²Ñ‹Ð±Ð¾Ñ€Ð° ÑÑ€Ð¾ÐºÐ° Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸
delivery_keyboard = InlineKeyboardMarkup(
    inline_keyboard=[
        [InlineKeyboardButton(text="10-13 Ð´Ð½ÐµÐ¹", callback_data="10-13")],
        [InlineKeyboardButton(text="15-18 Ð´Ð½ÐµÐ¹", callback_data="15-18")],
        [InlineKeyboardButton(text="25-30 Ð´Ð½ÐµÐ¹", callback_data="25-30")],
        [InlineKeyboardButton(text="ÐÐ²Ð¸Ð°", callback_data="avia")]
    ]
)

# ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ñ‡Ð¸Ðº ÐºÐ¾Ð¼Ð°Ð½Ð´Ñ‹ /start
@dp.message(Command("start"))
async def start(message: types.Message):
    """ÐžÑ‡Ð¸Ñ‰Ð°ÐµÑ‚ Ð´Ð°Ð½Ð½Ñ‹Ðµ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»Ñ Ð¸ Ð¿Ð¾ÐºÐ°Ð·Ñ‹Ð²Ð°ÐµÑ‚ Ð³Ð»Ð°Ð²Ð½Ð¾Ðµ Ð¼ÐµÐ½ÑŽ."""
    chat_id = message.chat.id
    if chat_id in user_answers:
        del user_answers[chat_id]
    for media_group_id, group_data in list(media_groups.items()):
        if group_data["chat_id"] == chat_id:
            del media_groups[media_group_id]
    await message.answer("ÐŸÑ€Ð¸Ð²ÐµÑ‚! Ð¯ Ð¿Ð¾Ð¼Ð¾Ð³Ñƒ Ð²Ð°Ð¼ Ñ Ð·Ð°ÐºÐ°Ð·Ð¾Ð¼. Ð’Ñ‹Ð±ÐµÑ€Ð¸Ñ‚Ðµ Ð´ÐµÐ¹ÑÑ‚Ð²Ð¸Ðµ:", reply_markup=start_keyboard)

# ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ñ‡Ð¸Ðº ÐºÐ½Ð¾Ð¿ÐºÐ¸ "Ð—Ð°Ð¿Ð¾Ð»Ð½Ð¸Ñ‚ÑŒ Ð°Ð½ÐºÐµÑ‚Ñƒ"
@dp.message(lambda message: message.text == "Ð—Ð°Ð¿Ð¾Ð»Ð½Ð¸Ñ‚ÑŒ Ð°Ð½ÐºÐµÑ‚Ñƒ")
async def request_consent(message: types.Message):
    """Ð—Ð°Ð¿Ñ€Ð°ÑˆÐ¸Ð²Ð°ÐµÑ‚ ÑÐ¾Ð³Ð»Ð°ÑÐ¸Ðµ Ð½Ð° Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÑƒ Ð¿ÐµÑ€ÑÐ¾Ð½Ð°Ð»ÑŒÐ½Ñ‹Ñ… Ð´Ð°Ð½Ð½Ñ‹Ñ…."""
    await message.answer(
        "ÐŸÐµÑ€ÐµÐ´ Ð·Ð°Ð¿Ð¾Ð»Ð½ÐµÐ½Ð¸ÐµÐ¼ Ð°Ð½ÐºÐµÑ‚Ñ‹, Ð¿Ð¾Ð¶Ð°Ð»ÑƒÐ¹ÑÑ‚Ð°, Ð´Ð°Ð¹Ñ‚Ðµ ÑÐ¾Ð³Ð»Ð°ÑÐ¸Ðµ Ð½Ð° Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÑƒ Ð¿ÐµÑ€ÑÐ¾Ð½Ð°Ð»ÑŒÐ½Ñ‹Ñ… Ð´Ð°Ð½Ð½Ñ‹Ñ….",
        reply_markup=consent_keyboard
    )

# ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ñ‡Ð¸Ðº callback-Ð·Ð°Ð¿Ñ€Ð¾ÑÐ¾Ð² Ð´Ð»Ñ ÑÐ¾Ð³Ð»Ð°ÑÐ¸Ñ
@dp.callback_query(lambda call: call.data in ["consent_yes", "consent_no", "view_offer", "final_no"])
async def process_consent(call: types.CallbackQuery):
    """ÐžÐ±Ñ€Ð°Ð±Ð°Ñ‚Ñ‹Ð²Ð°ÐµÑ‚ Ð²Ñ‹Ð±Ð¾Ñ€ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»Ñ Ð¿Ð¾ ÑÐ¾Ð³Ð»Ð°ÑÐ¸ÑŽ Ð½Ð° Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÑƒ Ð´Ð°Ð½Ð½Ñ‹Ñ…."""
    chat_id = call.message.chat.id
    if call.data == "consent_yes":
        global survey_id_counter
        user_answers[chat_id] = {"id": survey_id_counter, "answers": [], "source_chat": call.message.chat.id}
        survey_id_counter += 1
        save_survey_id(survey_id_counter)
        if str(chat_id) in users:
            user_answers[chat_id]["answers"] = [
                users[str(chat_id)]["name"],
                users[str(chat_id)]["telegram"],
                users[str(chat_id)]["phone"]
            ]
            await call.message.edit_text(f"Ð¡Ð¿Ð°ÑÐ¸Ð±Ð¾ Ð·Ð° ÑÐ¾Ð³Ð»Ð°ÑÐ¸Ðµ! ðŸ“ Ð’Ñ‹ ÑƒÐ¶Ðµ Ð·Ð°Ð¿Ð¾Ð»Ð½ÑÐ»Ð¸ ÐºÐ¾Ð½Ñ‚Ð°ÐºÑ‚Ð½Ñ‹Ðµ Ð´Ð°Ð½Ð½Ñ‹Ðµ, Ð½Ð°Ñ‡Ð½ÐµÐ¼ Ñ Ñ‚Ð¾Ð²Ð°Ñ€Ð°.\n\n{questions[3]}")
        else:
            await call.message.edit_text(f"Ð¡Ð¿Ð°ÑÐ¸Ð±Ð¾ Ð·Ð° ÑÐ¾Ð³Ð»Ð°ÑÐ¸Ðµ! ðŸ“ ÐÐ°Ñ‡Ð½ÐµÐ¼.\n\n{questions[0]}")
    elif call.data == "view_offer":
        try:
            await bot.send_document(chat_id, document=types.FSInputFile("/home/anna/syncli_bot/offer.pdf"), caption="ðŸ“„ ÐžÑ„ÐµÑ€Ñ‚Ð°")
            await call.message.edit_text("ÐžÐ·Ð½Ð°ÐºÐ¾Ð¼ÑŒÑ‚ÐµÑÑŒ Ñ Ð´Ð¾ÐºÑƒÐ¼ÐµÐ½Ñ‚Ð¾Ð¼ Ð¸ Ð²Ñ‹Ð±ÐµÑ€Ð¸Ñ‚Ðµ Ð²Ð°Ñ€Ð¸Ð°Ð½Ñ‚.", reply_markup=consent_keyboard)
        except Exception as e:
            logger.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ñ€Ð¸ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐµ Ð¾Ñ„ÐµÑ€Ñ‚Ñ‹: {str(e)}")
            await call.message.edit_text("ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ñ€Ð¸ Ð·Ð°Ð³Ñ€ÑƒÐ·ÐºÐµ Ð¾Ñ„ÐµÑ€Ñ‚Ñ‹. ÐŸÐ¾Ð¿Ñ€Ð¾Ð±ÑƒÐ¹Ñ‚Ðµ Ð¿Ð¾Ð·Ð¶Ðµ.", reply_markup=consent_keyboard)
    elif call.data == "consent_no":
        await call.message.edit_text(
            "ÐœÑ‹ Ð½Ðµ Ð¿ÐµÑ€ÐµÐ´Ð°ÐµÐ¼ Ð²Ð°ÑˆÐ¸ Ð´Ð°Ð½Ð½Ñ‹Ðµ Ñ‚Ñ€ÐµÑ‚ÑŒÐ¸Ð¼ Ð»Ð¸Ñ†Ð°Ð¼. ÐžÐ½Ð¸ Ð½ÑƒÐ¶Ð½Ñ‹ Ñ‚Ð¾Ð»ÑŒÐºÐ¾ Ð´Ð»Ñ Ð·Ð°ÐºÐ°Ð·Ð°. ÐŸÑ€Ð¾Ð´Ð¾Ð»Ð¶Ð¸Ñ‚ÑŒ?",
            reply_markup=consent_second_keyboard
        )
    elif call.data == "final_no":
        await call.message.edit_text("Ð¡Ð²ÑÐ¶Ð¸Ñ‚ÐµÑÑŒ Ñ Ð¼ÐµÐ½ÐµÐ´Ð¶ÐµÑ€Ð¾Ð¼ Ð½Ð°Ð¿Ñ€ÑÐ¼ÑƒÑŽ: @YourManagerTelegram", reply_markup=None)

# ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ñ‡Ð¸Ðº Ñ„Ð°Ð¹Ð»Ð¾Ð² (Ñ„Ð¾Ñ‚Ð¾/Ð´Ð¾ÐºÑƒÐ¼ÐµÐ½Ñ‚Ñ‹)
@dp.message(lambda message: message.photo or message.document)
async def handle_file(message: types.Message):
    """ÐžÐ±Ñ€Ð°Ð±Ð°Ñ‚Ñ‹Ð²Ð°ÐµÑ‚ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÑƒ Ñ„Ð°Ð¹Ð»Ð¾Ð² Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»ÐµÐ¼ Ð½Ð° ÑÑ‚Ð°Ð¿Ðµ Ð°Ð½ÐºÐµÑ‚Ñ‹."""
    chat_id = message.chat.id
    logger.debug(f"ÐŸÐ¾Ð»ÑƒÑ‡ÐµÐ½Ð¾ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ. Chat ID: {chat_id}, Media Group ID: {message.media_group_id}, Photo: {bool(message.photo)}, Document: {bool(message.document)}")
    if chat_id in user_answers and len(user_answers[chat_id]["answers"]) == 6:
        if not any(isinstance(answer, list) for answer in user_answers[chat_id]["answers"]):
            user_answers[chat_id]["answers"].append([])
            logger.debug(f"Ð¡Ð¾Ð·Ð´Ð°Ð½ Ð½Ð¾Ð²Ñ‹Ð¹ ÑÐ¿Ð¸ÑÐ¾Ðº Ñ„Ð°Ð¹Ð»Ð¾Ð² Ð´Ð»Ñ Chat ID: {chat_id}")
        files_list = user_answers[chat_id]["answers"][6]
        if message.media_group_id:
            if message.media_group_id not in media_groups:
                media_groups[message.media_group_id] = {"chat_id": chat_id, "files": [], "processed": False}
            file_data = None
            if message.photo:
                photo = message.photo[-1]
                file_data = {"file_id": photo.file_id, "type": "photo"}
            elif message.document:
                file_data = {"file_id": message.document.file_id, "type": "document"}
            if file_data and file_data["file_id"] not in [f["file_id"] for f in media_groups[message.media_group_id]["files"]]:
                media_groups[message.media_group_id]["files"].append(file_data)
                logger.debug(f"Ð”Ð¾Ð±Ð°Ð²Ð»ÐµÐ½ Ñ„Ð°Ð¹Ð» Ð² Ð¼ÐµÐ´Ð¸Ð°Ð³Ñ€ÑƒÐ¿Ð¿Ñƒ {message.media_group_id}: {file_data['file_id']}")
            if not media_groups[message.media_group_id]["processed"]:
                media_groups[message.media_group_id]["processed"] = True
                await message.answer("âœ… Ð¤Ð°Ð¹Ð»(Ñ‹) Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ñ‹. ÐŸÑ€Ð¸ÐºÑ€ÐµÐ¿Ð¸Ñ‚Ðµ ÐµÑ‰Ðµ Ð¸Ð»Ð¸ Ð½Ð°Ð¿Ð¸ÑˆÐ¸Ñ‚Ðµ 'Ð“Ð¾Ñ‚Ð¾Ð²Ð¾' Ð´Ð»Ñ Ð¿Ñ€Ð¾Ð´Ð¾Ð»Ð¶ÐµÐ½Ð¸Ñ.")
            return
        else:
            if message.photo:
                photo = message.photo[-1]
                files_list.append({"file_id": photo.file_id, "type": "photo"})
                logger.debug(f"Ð”Ð¾Ð±Ð°Ð²Ð»ÐµÐ½Ð¾ Ð¾Ð´Ð¸Ð½Ð¾Ñ‡Ð½Ð¾Ðµ Ñ„Ð¾Ñ‚Ð¾: {photo.file_id}")
            elif message.document:
                files_list.append({"file_id": message.document.file_id, "type": "document"})
                logger.debug(f"Ð”Ð¾Ð±Ð°Ð²Ð»ÐµÐ½ Ð¾Ð´Ð¸Ð½Ð¾Ñ‡Ð½Ñ‹Ð¹ Ð´Ð¾ÐºÑƒÐ¼ÐµÐ½Ñ‚: {message.document.file_id}")
            await message.answer("âœ… Ð¤Ð°Ð¹Ð» Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½. ÐŸÑ€Ð¸ÐºÑ€ÐµÐ¿Ð¸Ñ‚Ðµ ÐµÑ‰Ðµ Ð¸Ð»Ð¸ Ð½Ð°Ð¿Ð¸ÑˆÐ¸Ñ‚Ðµ 'Ð“Ð¾Ñ‚Ð¾Ð²Ð¾' Ð´Ð»Ñ Ð¿Ñ€Ð¾Ð´Ð¾Ð»Ð¶ÐµÐ½Ð¸Ñ.")
            logger.debug(f"Ð¢ÐµÐºÑƒÑ‰Ð¸Ð¹ ÑÐ¿Ð¸ÑÐ¾Ðº Ñ„Ð°Ð¹Ð»Ð¾Ð² Ð´Ð»Ñ Chat ID {chat_id}: {files_list}")
        return
    await message.answer("ðŸ“Ž ÐžÑ‚Ð¿Ñ€Ð°Ð²ÑŒÑ‚Ðµ Ñ„Ð°Ð¹Ð» Ñ‚Ð¾Ð»ÑŒÐºÐ¾ Ð½Ð° ÑÑ‚Ð°Ð¿Ðµ ÑÐ¾Ð¾Ñ‚Ð²ÐµÑ‚ÑÑ‚Ð²ÑƒÑŽÑ‰ÐµÐ³Ð¾ Ð²Ð¾Ð¿Ñ€Ð¾ÑÐ° Ð² Ð°Ð½ÐºÐµÑ‚Ðµ.")

# ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ñ‡Ð¸Ðº ÐºÐ¾Ð¼Ð°Ð½Ð´Ñ‹ "Ð“Ð¾Ñ‚Ð¾Ð²Ð¾" Ð´Ð»Ñ Ð·Ð°Ð²ÐµÑ€ÑˆÐµÐ½Ð¸Ñ Ð·Ð°Ð³Ñ€ÑƒÐ·ÐºÐ¸ Ñ„Ð°Ð¹Ð»Ð¾Ð²
@dp.message(lambda message: message.text and message.text.lower() == "Ð³Ð¾Ñ‚Ð¾Ð²Ð¾")
async def handle_ready(message: types.Message):
    """Ð—Ð°Ð²ÐµÑ€ÑˆÐ°ÐµÑ‚ ÑÑ‚Ð°Ð¿ Ð·Ð°Ð³Ñ€ÑƒÐ·ÐºÐ¸ Ñ„Ð°Ð¹Ð»Ð¾Ð² Ð¸ Ð¿ÐµÑ€ÐµÑ…Ð¾Ð´Ð¸Ñ‚ Ðº ÑÐ»ÐµÐ´ÑƒÑŽÑ‰ÐµÐ¼Ñƒ Ð²Ð¾Ð¿Ñ€Ð¾ÑÑƒ."""
    chat_id = message.chat.id
    logger.debug(f"ÐŸÐ¾Ð»ÑƒÑ‡ÐµÐ½Ð¾ 'Ð“Ð¾Ñ‚Ð¾Ð²Ð¾'. Chat ID: {chat_id}, user_answers: {user_answers.get(chat_id)}")
    if chat_id in user_answers:
        if len(user_answers[chat_id]["answers"]) >= 6 and isinstance(user_answers[chat_id]["answers"][6], list):
            files_list = user_answers[chat_id]["answers"][6]
            for media_group_id, group_data in list(media_groups.items()):
                if group_data["chat_id"] == chat_id:
                    files_list.extend([f for f in group_data["files"] if f["file_id"] not in [x["file_id"] for x in files_list]])
                    del media_groups[media_group_id]
                    logger.debug(f"ÐŸÐµÑ€ÐµÐ½ÐµÑÐµÐ½Ñ‹ Ñ„Ð°Ð¹Ð»Ñ‹ Ð¸Ð· Ð¼ÐµÐ´Ð¸Ð°Ð³Ñ€ÑƒÐ¿Ð¿Ñ‹ {media_group_id} Ð² ÑÐ¿Ð¸ÑÐ¾Ðº Chat ID {chat_id}: {files_list}")
            if len(user_answers[chat_id]["answers"]) == 7:
                await message.answer(questions[7])
            else:
                user_answers[chat_id]["answers"].append("Ð“Ð¾Ñ‚Ð¾Ð²Ð¾")
                await message.answer("âœ… Ð’ÑÐµ Ñ„Ð°Ð¹Ð»Ñ‹ Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ñ‹. ÐŸÑ€Ð¾Ð´Ð¾Ð»Ð¶Ð°ÐµÐ¼ Ð·Ð°Ð¿Ð¾Ð»Ð½ÐµÐ½Ð¸Ðµ Ð°Ð½ÐºÐµÑ‚Ñ‹.")
                await message.answer(questions[7])
            logger.debug(f"ÐŸÐ¾ÑÐ»Ðµ 'Ð“Ð¾Ñ‚Ð¾Ð²Ð¾' user_answers Ð´Ð»Ñ Chat ID {chat_id}: {user_answers[chat_id]}")
            return

# ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ñ‡Ð¸Ðº ÐºÐ½Ð¾Ð¿ÐºÐ¸ "Ð§Ð°ÑÑ‚Ñ‹Ðµ Ð²Ð¾Ð¿Ñ€Ð¾ÑÑ‹"
@dp.message(lambda message: message.text == "Ð§Ð°ÑÑ‚Ñ‹Ðµ Ð²Ð¾Ð¿Ñ€Ð¾ÑÑ‹")
async def show_faq(message: types.Message):
    """ÐŸÐ¾ÐºÐ°Ð·Ñ‹Ð²Ð°ÐµÑ‚ ÑÐ¿Ð¸ÑÐ¾Ðº Ñ‡Ð°ÑÑ‚Ð¾ Ð·Ð°Ð´Ð°Ð²Ð°ÐµÐ¼Ñ‹Ñ… Ð²Ð¾Ð¿Ñ€Ð¾ÑÐ¾Ð² Ð¸Ð· ÑÐ»Ð¾Ð²Ð°Ñ€Ñ faq, Ñ€Ð°Ð·Ð±Ð¸Ð²Ð°Ñ Ð½Ð° Ñ‡Ð°ÑÑ‚Ð¸."""
    chat_id = message.chat.id
    if chat_id in user_answers:
        del user_answers[chat_id]
    
    logger.debug(f"FAQ Ð´Ð»Ñ Ð¾Ñ‚Ð¾Ð±Ñ€Ð°Ð¶ÐµÐ½Ð¸Ñ: {faq}")
    if not faq:
        await message.answer("Ð¡Ð¿Ð¸ÑÐ¾Ðº Ð²Ð¾Ð¿Ñ€Ð¾ÑÐ¾Ð² Ð²Ñ€ÐµÐ¼ÐµÐ½Ð½Ð¾ Ð½ÐµÐ´Ð¾ÑÑ‚ÑƒÐ¿ÐµÐ½.")
        return

    # Ð¤Ð¾Ñ€Ð¼Ð¸Ñ€ÑƒÐµÐ¼ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ
    base_text = "ðŸ“Œ Ð§Ð°ÑÑ‚Ð¾ Ð·Ð°Ð´Ð°Ð²Ð°ÐµÐ¼Ñ‹Ðµ Ð²Ð¾Ð¿Ñ€Ð¾ÑÑ‹:\n\n"
    messages = []
    current_message = base_text
    max_length = 3900  # Ð£Ð¼ÐµÐ½ÑŒÑˆÐµÐ½Ð½Ñ‹Ð¹ Ð»Ð¸Ð¼Ð¸Ñ‚

    for question in faq.keys():
        line = f"ðŸ‘‰ {question.capitalize()}\n"
        if len(current_message) + len(line) > max_length:
            messages.append(current_message.strip())
            current_message = base_text
        current_message += line
    
    current_message += "\nÐÐ°Ð¿Ð¸ÑˆÐ¸Ñ‚Ðµ Ð²Ð°Ñˆ Ð²Ð¾Ð¿Ñ€Ð¾Ñ!"
    messages.append(current_message.strip())

    # ÐŸÑ€Ð¾Ð²ÐµÑ€ÑÐµÐ¼ Ð´Ð»Ð¸Ð½Ñƒ Ð¸ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²Ð»ÑÐµÐ¼
    for i, msg in enumerate(messages):
        logger.debug(f"Ð§Ð°ÑÑ‚ÑŒ {i+1}/{len(messages)}, Ð´Ð»Ð¸Ð½Ð°: {len(msg)}")
        if len(msg) > 4096:
            logger.error(f"Ð¡Ð¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ ÑÐ»Ð¸ÑˆÐºÐ¾Ð¼ Ð´Ð»Ð¸Ð½Ð½Ð¾Ðµ: {len(msg)} ÑÐ¸Ð¼Ð²Ð¾Ð»Ð¾Ð²")
            await message.answer("ÐžÑˆÐ¸Ð±ÐºÐ°: ÑÐ¿Ð¸ÑÐ¾Ðº Ð²Ð¾Ð¿Ñ€Ð¾ÑÐ¾Ð² ÑÐ»Ð¸ÑˆÐºÐ¾Ð¼ Ð´Ð»Ð¸Ð½Ð½Ñ‹Ð¹.")
            return
        try:
            await message.answer(msg)
            logger.debug(f"Ð£ÑÐ¿ÐµÑˆÐ½Ð¾ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð° Ñ‡Ð°ÑÑ‚ÑŒ {i+1}")
        except Exception as e:
            logger.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐ¸: {str(e)}")
            await message.answer("ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ñ€Ð¸ Ð¾Ñ‚Ð¾Ð±Ñ€Ð°Ð¶ÐµÐ½Ð¸Ð¸ Ð²Ð¾Ð¿Ñ€Ð¾ÑÐ¾Ð².")
            return

# Ð¤ÑƒÐ½ÐºÑ†Ð¸Ñ Ð·Ð°Ð²ÐµÑ€ÑˆÐµÐ½Ð¸Ñ Ð°Ð½ÐºÐµÑ‚Ñ‹ Ð¸ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐ¸ Ð´Ð°Ð½Ð½Ñ‹Ñ… Ð¼ÐµÐ½ÐµÐ´Ð¶ÐµÑ€Ñƒ
async def finish_survey(chat_id, message):
    """Ð¡Ð¾Ñ…Ñ€Ð°Ð½ÑÐµÑ‚ Ð´Ð°Ð½Ð½Ñ‹Ðµ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»Ñ Ð¸ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²Ð»ÑÐµÑ‚ Ð°Ð½ÐºÐµÑ‚Ñƒ Ð¼ÐµÐ½ÐµÐ´Ð¶ÐµÑ€Ñƒ."""
    if str(chat_id) not in users and len(user_answers[chat_id]["answers"]) >= 3:
        users[str(chat_id)] = {
            "name": user_answers[chat_id]["answers"][0],
            "telegram": user_answers[chat_id]["answers"][1],
            "phone": user_answers[chat_id]["answers"][2]
        }
        save_users(users)
    answers_text = "\n".join(f"{questions[i]}: {answer}" for i, answer in enumerate(user_answers[chat_id]["answers"]) if not isinstance(answer, list))
    try:
        files = user_answers[chat_id]["answers"][6] if len(user_answers[chat_id]["answers"]) > 6 and isinstance(user_answers[chat_id]["answers"][6], list) else []
        logger.debug(f"ÐžÑ‚Ð¿Ñ€Ð°Ð²ÐºÐ° Ñ„Ð°Ð¹Ð»Ð¾Ð² Ð¼ÐµÐ½ÐµÐ´Ð¶ÐµÑ€Ñƒ Ð´Ð»Ñ Chat ID {chat_id}: {files}")
        if files:
            media_group = []
            unique_file_ids = set()
            for file in files:
                file_id = file["file_id"]
                if file_id not in unique_file_ids:
                    if file["type"] == "photo":
                        media_group.append(InputMediaPhoto(media=file_id))
                    elif file["type"] == "document":
                        media_group.append(InputMediaDocument(media=file_id))
                    unique_file_ids.add(file_id)
            if media_group:
                media_group[0].caption = f"ðŸ“© ÐÐ¾Ð²Ð°Ñ Ð°Ð½ÐºÐµÑ‚Ð° ID {user_answers[chat_id]['id']}:\n\n{answers_text}"
                await bot.send_media_group(MANAGER_CHAT_ID, media=media_group)
                logger.debug(f"ÐžÑ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð° Ð¼ÐµÐ´Ð¸Ð°Ð³Ñ€ÑƒÐ¿Ð¿Ð° Ñ {len(media_group)} Ñ„Ð°Ð¹Ð»Ð°Ð¼Ð¸")
        else:
            await bot.send_message(MANAGER_CHAT_ID, f"ðŸ“© ÐÐ¾Ð²Ð°Ñ Ð°Ð½ÐºÐµÑ‚Ð° ID {user_answers[chat_id]['id']}:\n\n{answers_text}")
        await message.answer("Ð’Ð°ÑˆÐ° Ð°Ð½ÐºÐµÑ‚Ð° ÑƒÑÐ¿ÐµÑˆÐ½Ð¾ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð°! ÐœÑ‹ ÑÐ²ÑÐ¶ÐµÐ¼ÑÑ Ñ Ð²Ð°Ð¼Ð¸ Ð² Ð±Ð»Ð¸Ð¶Ð°Ð¹ÑˆÐµÐµ Ð²Ñ€ÐµÐ¼Ñ.")
    except Exception as e:
        logger.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ñ€Ð¸ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐµ: {e}")
        await message.answer("ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ñ€Ð¸ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐµ Ð°Ð½ÐºÐµÑ‚Ñ‹. Ð¡Ð²ÑÐ¶Ð¸Ñ‚ÐµÑÑŒ Ñ Ð¼ÐµÐ½ÐµÐ´Ð¶ÐµÑ€Ð¾Ð¼: @YourManagerTelegram")
    for media_group_id, group_data in list(media_groups.items()):
        if group_data["chat_id"] == chat_id:
            del media_groups[media_group_id]
    del user_answers[chat_id]

# ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ñ‡Ð¸Ðº Ñ‚ÐµÐºÑÑ‚Ð¾Ð²Ñ‹Ñ… ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹ Ð´Ð»Ñ Ð°Ð½ÐºÐµÑ‚Ñ‹ Ð¸ FAQ
@dp.message(lambda message: message.text)
async def collect_answers_or_faq(message: types.Message):
    """Ð¡Ð¾Ð±Ð¸Ñ€Ð°ÐµÑ‚ Ð¾Ñ‚Ð²ÐµÑ‚Ñ‹ Ð½Ð° Ð°Ð½ÐºÐµÑ‚Ñƒ Ð¸Ð»Ð¸ Ð¾Ñ‚Ð²ÐµÑ‡Ð°ÐµÑ‚ Ð½Ð° Ð²Ð¾Ð¿Ñ€Ð¾ÑÑ‹ Ð¸Ð· FAQ Ñ ÑƒÐ»ÑƒÑ‡ÑˆÐµÐ½Ð½Ñ‹Ð¼ Ð¿Ð¾Ð¸ÑÐºÐ¾Ð¼."""
    chat_id = message.chat.id
    text = message.text.lower()

    # ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÐ° ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹ Ð² Ð³Ñ€ÑƒÐ¿Ð¿Ð°Ñ… (Ñ‚Ð¾Ð»ÑŒÐºÐ¾ FAQ)
    if message.chat.type in [ChatType.GROUP, ChatType.SUPERGROUP]:
        best_match = None
        max_words_matched = 0
        text_words = set(text.split())  # Ð Ð°Ð·Ð±Ð¸Ð²Ð°ÐµÐ¼ Ð·Ð°Ð¿Ñ€Ð¾Ñ Ð½Ð° ÑÐ»Ð¾Ð²Ð°

        for keyword, response in faq.items():
            keyword_words = set(keyword.lower().split())
            matched_words = len(text_words & keyword_words)  # Ð¡Ñ‡Ð¸Ñ‚Ð°ÐµÐ¼ ÑÐ¾Ð²Ð¿Ð°Ð´ÐµÐ½Ð¸Ñ ÑÐ»Ð¾Ð²
            if matched_words > max_words_matched:
                max_words_matched = matched_words
                best_match = response
        if best_match:
            await message.reply(best_match)
        return

    # ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÐ° FAQ Ð² Ð»Ð¸Ñ‡Ð½Ñ‹Ñ… Ñ‡Ð°Ñ‚Ð°Ñ…
    best_match = None
    max_words_matched = 0
    text_words = set(text.split())  # Ð Ð°Ð·Ð±Ð¸Ð²Ð°ÐµÐ¼ Ð·Ð°Ð¿Ñ€Ð¾Ñ Ð½Ð° ÑÐ»Ð¾Ð²Ð°

    for keyword, response in faq.items():
        keyword_words = set(keyword.lower().split())
        matched_words = len(text_words & keyword_words)  # Ð¡Ñ‡Ð¸Ñ‚Ð°ÐµÐ¼ ÑÐ¾Ð²Ð¿Ð°Ð´ÐµÐ½Ð¸Ñ ÑÐ»Ð¾Ð²
        if matched_words > max_words_matched:
            max_words_matched = matched_words
            best_match = response
    if best_match:
        await message.answer(best_match)
        return

    # ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÐ° Ð°Ð½ÐºÐµÑ‚Ñ‹
    if chat_id in user_answers:
        if text == "Ð½Ð°Ð·Ð°Ð´" and user_answers[chat_id]["answers"]:
            user_answers[chat_id]["answers"].pop()
            await message.answer(f"ðŸ”„ Ð’Ð²ÐµÐ´Ð¸Ñ‚Ðµ Ð½Ð¾Ð²Ñ‹Ð¹ Ð¾Ñ‚Ð²ÐµÑ‚:\n\n{questions[len(user_answers[chat_id]['answers'])]}")
            return
        if text == "Ð³Ð¾Ñ‚Ð¾Ð²Ð¾":
            await handle_ready(message)
            return
        user_answers[chat_id]["answers"].append(message.text)
        next_index = len(user_answers[chat_id]["answers"])
        if next_index < len(questions):
            if next_index == 12:
                await message.answer("â³ Ð’Ñ‹Ð±ÐµÑ€Ð¸Ñ‚Ðµ ÑÑ€Ð¾Ðº Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸:", reply_markup=delivery_keyboard)
            else:
                await message.answer(questions[next_index])
        else:
            await finish_survey(chat_id, message)
    else:
        if message.chat.type == ChatType.PRIVATE:
            await message.answer("Ð¯ Ð¿Ð¾ÐºÐ° Ð½Ðµ Ð·Ð½Ð°ÑŽ Ð¾Ñ‚Ð²ÐµÑ‚Ð° Ð½Ð° ÑÑ‚Ð¾Ñ‚ Ð²Ð¾Ð¿Ñ€Ð¾Ñ, Ð½Ð¾ Ð¿ÐµÑ€ÐµÐ´Ð°Ð¼ ÐµÐ³Ð¾ Ð¼ÐµÐ½ÐµÐ´Ð¶ÐµÑ€Ñƒ!")

# ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ñ‡Ð¸Ðº Ð²Ñ‹Ð±Ð¾Ñ€Ð° ÑÑ€Ð¾ÐºÐ° Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸
@dp.callback_query(lambda call: call.data in ["10-13", "15-18", "25-30", "avia"])
async def delivery_selected(call: types.CallbackQuery):
    """ÐžÐ±Ñ€Ð°Ð±Ð°Ñ‚Ñ‹Ð²Ð°ÐµÑ‚ Ð²Ñ‹Ð±Ð¾Ñ€ ÑÑ€Ð¾ÐºÐ° Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸ Ð¸Ð· Ð¸Ð½Ð»Ð°Ð¹Ð½-ÐºÐ»Ð°Ð²Ð¸Ð°Ñ‚ÑƒÑ€Ñ‹."""
    chat_id = call.message.chat.id
    if chat_id in user_answers:
        user_answers[chat_id]["answers"].append(call.data)
        next_index = len(user_answers[chat_id]["answers"])
        if next_index < len(questions):
            await call.message.answer(f"âœ… Ð¡Ñ€Ð¾Ðº Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸ Ð²Ñ‹Ð±Ñ€Ð°Ð½.\n\n{questions[next_index]}")
        else:
            await finish_survey(chat_id, call.message)
    await call.answer()

# Ð“Ð»Ð°Ð²Ð½Ð°Ñ Ñ„ÑƒÐ½ÐºÑ†Ð¸Ñ Ð·Ð°Ð¿ÑƒÑÐºÐ° Ð±Ð¾Ñ‚Ð°
async def main():
    """Ð—Ð°Ð¿ÑƒÑÐºÐ°ÐµÑ‚ Ð±Ð¾Ñ‚Ð°: Ð·Ð°Ð³Ñ€ÑƒÐ¶Ð°ÐµÑ‚ FAQ, Ð·Ð°Ð¿ÑƒÑÐºÐ°ÐµÑ‚ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð¸Ðµ Ð¸ Ð½Ð°Ñ‡Ð¸Ð½Ð°ÐµÑ‚ Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÑƒ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹."""
    try:
        logger.info("Ð—Ð°Ð¿ÑƒÑÐº Ð±Ð¾Ñ‚Ð°...")
        await update_faq()  # Ð—Ð°Ð³Ñ€ÑƒÐ¶Ð°ÐµÐ¼ FAQ Ð¿ÐµÑ€ÐµÐ´ Ð·Ð°Ð¿ÑƒÑÐºÐ¾Ð¼
        asyncio.create_task(periodic_faq_update())  # Ð—Ð°Ð¿ÑƒÑÐºÐ°ÐµÐ¼ Ð¿ÐµÑ€Ð¸Ð¾Ð´Ð¸Ñ‡ÐµÑÐºÐ¾Ðµ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð¸Ðµ FAQ
        logger.info("ÐÐ°Ñ‡Ð¸Ð½Ð°ÐµÐ¼ polling...")
        await dp.start_polling(bot)  # Ð—Ð°Ð¿ÑƒÑÐºÐ°ÐµÐ¼ Ð¾Ð¿Ñ€Ð¾Ñ Telegram API
    except Exception as e:
        logger.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ñ€Ð¸ Ð·Ð°Ð¿ÑƒÑÐºÐµ Ð±Ð¾Ñ‚Ð°: {str(e)}")
        raise  # ÐŸÐ¾Ð²Ñ‚Ð¾Ñ€Ð½Ð¾ Ð¿Ð¾Ð´Ð½Ð¸Ð¼Ð°ÐµÐ¼ Ð¸ÑÐºÐ»ÑŽÑ‡ÐµÐ½Ð¸Ðµ Ð´Ð»Ñ Ð´Ð¸Ð°Ð³Ð½Ð¾ÑÑ‚Ð¸ÐºÐ¸

if __name__ == "__main__":
    asyncio.run(main())